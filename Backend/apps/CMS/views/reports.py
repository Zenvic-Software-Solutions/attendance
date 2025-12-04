from rest_framework.views import APIView
from django.http import HttpResponse
from django.db.models import Sum
from datetime import date, datetime
import openpyxl
from openpyxl.utils import get_column_letter

from apps.CMS.models import Customer, Loan, Payment, Centre
from apps.CMS.serializers import (
    DailyReportSerializer,
    CenterReportSerializer,
    LoanReportSerializer
)


class ReportsAPIView(APIView):

    def get(self, request):
        report_type = request.GET.get("type")

        if report_type == "daily":
            data = self.daily_report_data(request)
        elif report_type == "center":
            data = self.center_report_data(request)
        elif report_type == "loan":
            data = self.loan_report_data(request)
        elif report_type == "loan_status":
            data = self.loan_status_data()
        elif report_type == "collection_summary":
            data = self.collection_summary_data(request)
        elif report_type == "center_outstanding":
            data = self.center_outstanding_data(request)
        elif report_type == "due_report":
            data = self.due_report_data(request)
        else:
            return HttpResponse("Invalid report type", status=400)

        return self.generate_excel(report_type, data)

    # -----------------------------------------------------
    # EXCEL GENERATOR
    # -----------------------------------------------------
    def generate_excel(self, title, data_list):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title.upper()

        if isinstance(data_list, dict):
            data_list = [data_list]

        if not data_list:
            data_list = [{}]

        headers = list(data_list[0].keys())
        ws.append(headers)

        for row in data_list:
            ws.append([row.get(h, "") for h in headers])

        for col_no, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_no)].width = 22

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=\"{title}.xlsx\"'
        wb.save(response)
        return response

    # -----------------------------------------------------
    # REPORT: DAILY REPORT (Filter: date)
    # -----------------------------------------------------
    def daily_report_data(self, request):
        date_str = request.GET.get("date")
        today = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else date.today()

        today_paid_total = Payment.objects.filter(
            status="Paid", paid_date=today
        ).aggregate(total=Sum("amount"))["total"] or 0

        today_paid_count = Payment.objects.filter(
            status="Paid", paid_date=today
        ).count()

        today_unpaid_count = Payment.objects.filter(status="Unpaid").count()

        loan_due_today = Loan.objects.filter(
            initiated_date__date=today
        ).count()

        data = {
            "date": today,
            "today_paid_total": today_paid_total,
            "today_paid_count": today_paid_count,
            "today_unpaid_count": today_unpaid_count,
            "loan_due_today": loan_due_today,
        }
        return DailyReportSerializer(data).data

    # -----------------------------------------------------
    # REPORT: CENTER SUMMARY (Filters: center_id, date)
    # -----------------------------------------------------
    def center_report_data(self, request):
        date_str = request.GET.get("date")
        filter_center = request.GET.get("center")
        today = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else date.today()

        result = []
        centers = Centre.objects.all()

        if filter_center:
            centers = centers.filter(id=filter_center)

        for center in centers:
            cust_ids = Customer.objects.filter(centre=center).values_list("id", flat=True)
            loans = Loan.objects.filter(customer_id__in=cust_ids)
            loan_ids = loans.values_list("id", flat=True)

            today_paid = Payment.objects.filter(
                loan_id__in=loan_ids, status="Paid", paid_date=today
            ).aggregate(total=Sum("amount"))["total"] or 0

            today_unpaid = Payment.objects.filter(
                loan_id__in=loan_ids, status="Unpaid"
            ).aggregate(total=Sum("amount"))["total"] or 0

            result.append({
                "center": center.identity,
                "date": today,
                "total_customers": len(cust_ids),
                "total_loans": loans.count(),
                "today_paid": today_paid,
                "today_unpaid": today_unpaid,
            })

        return result

    # -----------------------------------------------------
    # REPORT: LOAN REPORT (Filter: status=active/closed)
    # -----------------------------------------------------
    def loan_report_data(self, request):
        status = request.GET.get("status")  # active / closed
        loans = Loan.objects.all()

        if status == "active":
            loans = loans.filter(balance_week__gt=0)
        elif status == "closed":
            loans = loans.filter(balance_week=0)

        result = []

        for loan in loans:
            total_paid = Payment.objects.filter(
                loan=loan, status="Paid"
            ).aggregate(total=Sum("amount"))["total"] or 0

            result.append({
                "loan_title": loan.loan_title,
                "customer": loan.customer.identity,
                "principal_amount": loan.principal_amount,
                "interest_amount": loan.interest_amount,
                "total_paid": total_paid,
                "balance": loan.balance,
                "status": "Closed" if loan.balance_week == 0 else "Active"
            })

        return result

    # -----------------------------------------------------
    # REPORT: LOAN STATUS (no filters)
    # -----------------------------------------------------
    def loan_status_data(self):
        return {
            "total_loans": Loan.objects.count(),
            "active_loans": Loan.objects.filter(balance_week__gt=0).count(),
            "closed_loans": Loan.objects.filter(balance_week=0).count(),
        }

    # -----------------------------------------------------
    # REPORT: COLLECTION SUMMARY (Filters: start_date, end_date)
    # -----------------------------------------------------
    def collection_summary_data(self, request):
        start = request.GET.get("start")
        end = request.GET.get("end")

        qs = Payment.objects.filter(status__in=["Paid", "Unpaid"])

        if start:
            qs = qs.filter(paid_date__gte=start)

        if end:
            qs = qs.filter(paid_date__lte=end)

        total_paid = qs.filter(status="Paid").aggregate(total=Sum("amount"))["total"] or 0
        total_unpaid = qs.filter(status="Unpaid").aggregate(total=Sum("amount"))["total"] or 0

        return {
            "total_paid": total_paid,
            "total_unpaid": total_unpaid,
            "total_payments": total_paid + total_unpaid,
        }

    # -----------------------------------------------------
    # REPORT: CENTER OUTSTANDING (Filter: center_id)
    # -----------------------------------------------------
    def center_outstanding_data(self, request):
        filter_center = request.GET.get("center")

        centers = Centre.objects.all()
        if filter_center:
            centers = centers.filter(id=filter_center)

        result = []

        for center in centers:
            cust_ids = Customer.objects.filter(centre=center).values_list("id", flat=True)
            balance = Loan.objects.filter(customer_id__in=cust_ids).aggregate(total=Sum("balance"))["total"] or 0

            result.append({
                "center": center.identity,
                "outstanding_balance": balance,
            })

        return result

    # -----------------------------------------------------
    # REPORT: DUE REPORT (Filters: status, next_due_date)
    # -----------------------------------------------------
    def due_report_data(self, request):
        status = request.GET.get("status")   # active / closed
        due_date = request.GET.get("date")   # filter by next_due_date

        loans = Loan.objects.all()

        if status == "active":
            loans = loans.filter(balance_week__gt=0)
        elif status == "closed":
            loans = loans.filter(balance_week=0)

        result = []

        for loan in loans:
            next_due = loan.get_next_due_date()

            if due_date and str(next_due) != due_date:
                continue

            result.append({
                "loan_title": loan.loan_title,
                "customer": loan.customer.identity,
                "next_due_date": next_due,
                "status": "Closed" if loan.balance_week == 0 else "Active"
            })

        return result


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Q
from datetime import datetime
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from django.http import HttpResponse
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.db.models import Q

from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from django.http import HttpResponse
from datetime import datetime
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer
)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.db.models import Q
from datetime import datetime
from django.http import HttpResponse
from rest_framework.permissions import AllowAny
from rest_framework.views import APIView
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.db.models import Q

class LoanReportExportAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        user = request.user

        # ---------------- BASE QUERY ----------------
        qs = Loan.objects.filter(user=user)

        # ---------------- FILTER INPUTS ----------------
        search = request.GET.get("search")
        centre_id = request.GET.get("centre")

        # ---------------- APPLY SEARCH FILTER ----------------
        if search:
            qs = qs.filter(
                Q(customer__identity__icontains=search) |
                Q(customer__phone_number__icontains=search)
            )

        # ---------------- CENTRE FILTER ----------------
        if centre_id:
            qs = qs.filter(customer__centre__id=centre_id)

        # ---------------- PDF RESPONSE ----------------
        response = HttpResponse(content_type='application/pdf')
        filename = f"loan_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'

        # ---------------- LANDSCAPE PDF ----------------
        doc = SimpleDocTemplate(response, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()

        # ---------------- TITLE ----------------
        title = Paragraph("Loan Report", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 6))

        # ---------------- CURRENT DATE TOP RIGHT ----------------
        current_date = Paragraph(
            f"<para alignment='right'><b>Date:</b> {datetime.now().strftime('%Y-%m-%d')}</para>",
            styles['Normal']
        )
        elements.append(current_date)
        elements.append(Spacer(1, 12))

        # ---------------- HEADER ----------------
        data = [
            [
                "S.No","LoanDate", "Loan Title", "Customer", "Phone",
                "LoanAmount", "Centre", "Outstanding",
                "Total Week", "Paid Week", "EMI"
            ]
        ]

        wrap_style = ParagraphStyle(
            name="WrapStyle",
            fontSize=9,
            leading=11
        )

        # ---------------- TABLE DATA ----------------
        sno = 1
        for loan in qs:
            customer = loan.customer
            centre = customer.centre if customer else None

            balance_week = int(loan.balance_week or 0)
            total_week = int(loan.total_week or 0)
            paid_week = max(total_week - balance_week, 0)

            data.append([
                sno,
                loan.loan_date.strftime("%Y-%m-%d") if loan.loan_date else "",
                Paragraph(loan.loan_title or "", wrap_style),
                Paragraph(customer.identity if customer else "", wrap_style),
                Paragraph(customer.phone_number if customer else "", wrap_style),
                loan.principal_amount,
                Paragraph(centre.identity if centre else "", wrap_style),
                loan.balance,
                total_week,
                paid_week,
                loan.interest_amount,
            ])
            sno += 1

        # ---------------- EMI TOTAL CALCULATION ----------------
        emi_total = sum([(loan.interest_amount or 0) for loan in qs])

        # ---------------- ADD TOTAL EMI ROW ----------------
        emi_total_row = [
            "", "", "", "", "", "", "", "", "", "Total EMI", emi_total
        ]
        data.append(emi_total_row)

        # ---------------- TABLE FORMAT ----------------
        table = Table(
            data,
            colWidths=[40, 80, 110, 80, 70, 80, 70, 70, 60, 60, 70]
        )

        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTSIZE', (0, 0), (-1, 0), 10),

            # Body
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),

            # Alignment & wrapping
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),

            # Total EMI row style
            ('BACKGROUND', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
        ]))

        elements.append(table)

        # ---------------- BUILD PDF ----------------
        doc.build(elements)

        return response



class PaymentReportExportAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        try:
            # ---- FILTERS ----
            from_date = request.GET.get("from_date")
            to_date = request.GET.get("to_date")
            search = request.GET.get("search", "")
            user = request.user

            payments = Payment.objects.filter(user =user).select_related(
                "loan",
                "loan__customer",
                "loan__customer__centre"
            )

            # Apply date filter
            if from_date:
                payments = payments.filter(paid_date__gte=from_date)

            if to_date:
                payments = payments.filter(paid_date__lte=to_date)

            # Apply search filter
            if search:
                payments = payments.filter(
                    Q(loan__loan_title__icontains=search) |
                    Q(loan__customer__identity__icontains=search) |
                    Q(loan__customer__phone_number__icontains=search)
                )

            # --- Create Workbook ---
            wb = Workbook()
            ws = wb.active
            ws.title = "Payment Report"

            headers = [
                "S.No",
                "Loan Title",
                "Customer Name",
                "Customer Phone",
                "Centre",
                "Paid Amount",
                "Status",
                "Paid Date",
            ]
            ws.append(headers)
            sno = 1 

            # --- Fill rows ---
            for p in payments:
                ws.append([
                    
                    sno,
                    (p.loan.loan_title or "").strip(),
                    (p.loan.customer.identity or "").strip() if p.loan and p.loan.customer else "",
                    (p.loan.customer.phone_number or "").strip() if p.loan and p.loan.customer else "",
                    (p.loan.customer.centre.identity or "").strip() if p.loan and p.loan.customer.centre else "",
                    float(p.amount),
                    p.status.strip(),
                    p.paid_date.strftime("%Y-%m-%d") if p.paid_date else "",
                ])
                
                sno += 1 

            # --- Auto Column Width ---
            for idx, col in enumerate(ws.columns, 1):
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(idx)].width = max_length + 3

            # ---- Response ----
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = 'attachment; filename="payment_report.xlsx"'

            wb.save(response)
            return response

        except Exception as e:
            return Response({"error": str(e)}, status=500)



class TodayPaymentReportExportAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        try:
            search = request.GET.get("search", "")
            today = date.today()
            user = request.user

            # --- Query: Only today ---
            payments = Payment.objects.filter(user = user,paid_date=today).select_related(
                "loan",
                "loan__customer",
                "loan__customer__centre"
            )

            # --- Apply search only ---
            if search:
                payments = payments.filter(
                    Q(loan__loan_title__icontains=search) |
                    Q(loan__customer__identity__icontains=search) |
                    Q(loan__customer__phone_number__icontains=search)
                )

            # ---- Excel ----
            wb = Workbook()
            ws = wb.active
            ws.title = "Today Payment Report"

            headers = [
                "S.No",
                "Loan Title",
                "Customer Name",
                "Customer Phone",
                "Centre",
                "Paid Amount",
                "Status",
                "Paid Date",
            ]
            ws.append(headers)
            sno = 1 

            for p in payments:
                ws.append([
                    
                    sno,
                    (p.loan.loan_title or "").strip(),
                    (p.loan.customer.identity or "").strip() if p.loan and p.loan.customer else "",
                    (p.loan.customer.phone_number or "").strip() if p.loan and p.loan.customer else "",
                    (p.loan.customer.centre.identity or "").strip() if p.loan and p.loan.customer.centre else "",
                    float(p.amount),
                    p.status.strip(),
                    p.paid_date.strftime("%Y-%m-%d") if p.paid_date else "",
                ])
                
                sno += 1 

            # --- Auto size columns ---
            for idx, col in enumerate(ws.columns, 1):
                max_len = 0
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(idx)].width = max_len + 3

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = 'attachment; filename="today_payment_report.xlsx"'

            wb.save(response)
            return response

        except Exception as e:
            return Response({"error": str(e)}, status=500)

from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from django.http import HttpResponse

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


from django.db.models import Q

from datetime import date

class PaymentReportAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        try:
            # TODAY DATE FILTER
            today = date.today()
            user =request.user


            payments = Payment.objects.filter(user=user,
                paid_date=today
            ).select_related(
                "loan",
                "loan__customer",
                "loan__customer__centre"
            )

            # PDF RESPONSE
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="payment_report.pdf"'

            pdf = SimpleDocTemplate(response, pagesize=A4)
            elements = []

            styles = getSampleStyleSheet()
            title = Paragraph("<b>Payment Collection Report (Today)</b>", styles["Title"])
            elements.append(title)
            elements.append(Spacer(1, 20))

            # TABLE HEADER
            data = [
                ["S.No", "DATE", "CENTRE", "COLLECTION"]
            ]

            total_amount = 0
            sno = 1

            for p in payments:
                cust = p.loan.customer if p.loan else None
                centre_name = cust.centre.identity if cust and cust.centre else ""

                amount = float(p.amount)
                total_amount += amount

                row = [
                    sno,
                    p.paid_date.strftime("%d-%m-%Y") if p.paid_date else "",
                    centre_name,
                    amount
                ]
                data.append(row)
                sno += 1

            # TOTAL ROW
            data.append(["", "", "TOTAL", total_amount])

            table = Table(data, colWidths=[60, 120, 150, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))

            elements.append(table)
            pdf.build(elements)

            return response

        except Exception as e:
            return HttpResponse(str(e), status=500)
